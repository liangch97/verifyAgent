from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .utils import now_iso, write_json


def write_outputs(review: dict[str, Any], out_dir: Path, source_hash: str) -> dict[str, str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    md_path = out_dir / "contract_review.md"
    json_path = out_dir / f"contract_findings_{source_hash}.json"
    xlsx_path = out_dir / "合同形式化审核总表.xlsx"

    write_json(json_path, review)
    md_path.write_text(_render_markdown(review), encoding="utf-8")
    _write_excel_summary(xlsx_path, review, source_hash)

    return {
        "markdown": str(md_path),
        "json": str(json_path),
        "excel": str(xlsx_path),
    }


def _render_markdown(review: dict[str, Any]) -> str:
    extracted = review.get("extracted", {})
    findings = review.get("findings", [])
    not_checked = review.get("not_checked_items", [])
    external_checks = review.get("external_checks", {}) or {}
    summary_stats = review.get("summary_stats", {}) or {}

    risk_items = [f for f in findings if f.get("scope") == "enforce" and f.get("severity") in {"block", "warn"}]
    review_items = [f for f in findings if f.get("scope") == "review"]
    info_items = [f for f in findings if f.get("severity") == "info"]

    lines: list[str] = []
    lines.append("# 中山大学横向合同形式化审核报告")
    lines.append(f"- 生成时间：{review.get('generated_at', now_iso())}")
    lines.append(f"- 审核结论：**{review.get('decision', '-') }**")
    lines.append(f"- 置信度：{review.get('confidence', '-')}")
    lines.append("")

    lines.append("## 1. 行政摘要")
    lines.append(_render_admin_narrative(review, risk_items, review_items, info_items))
    lines.append("")
    lines.extend(_render_summary_stats(summary_stats))
    lines.append("")
    lines.append("### 1.1 建议先修改的事项")
    lines.extend(_render_admin_action_table(risk_items, empty_text="未发现必须先修改的合同文本问题。"))
    lines.append("")
    lines.append("### 1.2 需要人工确认或补充材料的事项")
    lines.extend(_render_admin_action_table(review_items, empty_text="未发现需要单独人工确认或补充材料的事项。"))
    lines.append("")
    if info_items:
        lines.append("### 1.3 信息提示")
        lines.extend(_render_admin_action_table(info_items, empty_text="无"))
        lines.append("")

    lines.append("## 2. 合同基本信息")
    lines.append(f"- 文件：`{review.get('source_file', '-')}`")
    lines.append(f"- 标题：{extracted.get('title', '-')}")
    lines.append(f"- 项目名称：{extracted.get('project_name', '-')}")
    lines.append(f"- 合同编号：{extracted.get('contract_no', '-')}")
    lines.append(f"- 甲方：{extracted.get('party_a', {}).get('name', '-')}")
    lines.append(f"- 乙方：{extracted.get('party_b', {}).get('name', '-')}")
    lines.append(f"- 合同类型：{extracted.get('contract_type', '-')}")
    lines.append(f"- 合同金额：{extracted.get('amount', '-')}")
    lines.append("")

    lines.append("## 外部核验材料")
    lines.extend(_render_external_checks(external_checks))
    lines.append("")

    lines.append("## 4. 风险项清单明细")
    lines.extend(_render_findings_table(risk_items))
    lines.append("")

    lines.append("## 5. 待人工复核项")
    lines.extend(_render_findings_table(review_items))
    lines.append("")

    lines.append("## 6. 信息提示项")
    lines.extend(_render_findings_table(info_items))
    lines.append("")

    lines.append("## 7. 未自动审核标准留档")
    if not not_checked:
        lines.append("- 无")
    else:
        lines.append("| 规则ID | 标题 | 说明 |")
        lines.append("|---|---|---|")
        for item in not_checked:
            lines.append(f"| {item.get('rule_id')} | {item.get('title')} | {item.get('message')} |")
    lines.append("")

    lines.append("## 8. 规则矩阵")
    lines.append("| 规则ID | 范围 | 严重度 | 标题 | 状态 |")
    lines.append("|---|---|---|---|---|")
    rule_results = review.get("rule_results") or []
    if rule_results:
        status_label = {"hit": "命中", "pass": "通过/未触发", "not_checked": "留档"}
        for item in rule_results:
            lines.append(
                f"| {item.get('rule_id')} | {item.get('scope')} | {item.get('severity')} | "
                f"{item.get('title')} | {status_label.get(item.get('status'), item.get('status'))} |"
            )
    else:
        all_rules = []
        all_rules.extend([(f.get("rule_id"), f.get("scope"), f.get("severity"), f.get("title"), "命中") for f in findings])
        all_rules.extend([(n.get("rule_id"), "not_checked", "not_checked", n.get("title"), "留档") for n in not_checked])
        for rid, scope, sev, title, hit in all_rules:
            lines.append(f"| {rid} | {scope} | {sev} | {title} | {hit} |")
    lines.append("")

    lines.append("## 9. 原文证据片段")
    if not findings:
        lines.append("- 未发现需展示证据的风险项。")
    for f in findings:
        lines.append(f"- `{f.get('rule_id')}` {f.get('title')}：{f.get('evidence')}（{f.get('location')}）")
    lines.append("")

    return "\n".join(lines)


def _render_summary_stats(summary_stats: dict[str, Any]) -> list[str]:
    if not summary_stats:
        return ["- 统计：本次未生成统计摘要，请重新运行审核。"]

    return [
        f"- 发现需处理事项：{summary_stats.get('findings_total', 0)} 项",
        f"- 建议先修改事项：{summary_stats.get('risk_total', 0)} 项",
        f"- 需要人工确认事项：{summary_stats.get('review_total', 0)} 项",
        f"- 提示事项：{summary_stats.get('info_total', 0)} 项",
        f"- 仅留档事项：{summary_stats.get('not_checked_total', 0)} 项",
    ]


def _render_admin_narrative(
    review: dict[str, Any],
    risk_items: list[dict[str, Any]],
    review_items: list[dict[str, Any]],
    info_items: list[dict[str, Any]],
) -> str:
    decision = review.get("decision", "-")
    parts = [f"本次审核结论为：**{decision}**。"]
    if risk_items:
        parts.append("建议先处理合同文本中可直接修改的事项，尤其是表格中列明的空白项、付款/账户、知识产权等条款。")
    if review_items:
        parts.append("另有部分事项需要项目负责人、合作方或经办老师补充材料后确认。")
    if not risk_items and not review_items and not info_items:
        parts.append("未发现需修改或补充确认的形式化问题。")
    parts.append("下表已把每个问题与对应条款或证据放在同一行，便于逐项核对和修改。")
    return "".join(parts)


def _format_rule_ids(rule_ids: list[Any]) -> str:
    return "、".join([str(rule_id) for rule_id in rule_ids]) if rule_ids else "-"


def _render_admin_action_table(findings: list[dict[str, Any]], empty_text: str) -> list[str]:
    if not findings:
        return [f"- {empty_text}"]

    out = ["| 序号 | 事项 | 关联条款/证据 | 风险说明 | 建议处理 |", "|---:|---|---|---|---|"]
    for idx, finding in enumerate(findings, start=1):
        title = _escape_md_cell(str(finding.get("title") or "需处理事项"))
        clause = _escape_md_cell(_format_clause_evidence(finding))
        impact = _escape_md_cell(_truncate_cell(str(finding.get("message") or "请人工复核"), 180))
        suggestion = _escape_md_cell(_truncate_cell(str(finding.get("suggestion") or "请根据审核要点补充或修订"), 200))
        out.append(f"| {idx} | {title} | {clause} | {impact} | {suggestion} |")
    return out


def _render_findings_table(findings: list[dict[str, Any]]) -> list[str]:
    if not findings:
        return ["- 无"]
    out = ["| 规则ID | 严重度 | 问题 | 合同位置 | 条款/证据 | 建议 |", "|---|---|---|---|---|---|"]
    for f in findings:
        issue = _escape_md_cell(str(f.get("message") or f.get("title") or "请人工复核"))
        location = _escape_md_cell(str(f.get("location") or "全文"))
        evidence = _escape_md_cell(_truncate_cell(str(f.get("evidence", "")), 320))
        sug = _escape_md_cell(_truncate_cell(str(f.get("suggestion", "")), 220))
        out.append(
            f"| {f.get('rule_id')} | {f.get('severity')} | {issue} | {location} | {evidence} | {sug} |"
        )
    return out


def _format_clause_evidence(finding: dict[str, Any]) -> str:
    location = str(finding.get("location") or "全文").strip()
    evidence = str(finding.get("evidence") or "未提取到直接证据").strip()
    if location and evidence:
        return _truncate_cell(f"{location}：{evidence}", 260)
    return _truncate_cell(evidence or location or "未提取到直接证据", 260)


def _escape_md_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\r", " ").replace("\n", " ").strip()


def _truncate_cell(value: str, limit: int) -> str:
    value = " ".join((value or "").split())
    if len(value) <= limit:
        return value
    return value[: max(0, limit - 3)].rstrip() + "..."


def _render_external_checks(external_checks: dict[str, Any]) -> list[str]:
    if not external_checks.get("company_check_provided"):
        return ["- 未提供外部工商核验材料；A1_REVIEW/A2/A3 保持待人工复核。"]

    party_external = (external_checks.get("party_a_match") or {}).get("external") or {}
    related = external_checks.get("related_party_check") or {}
    military = external_checks.get("military_defense_check") or {}
    evidence_files = external_checks.get("evidence_files") or []

    related_people = related.get("matched_people") or []
    related_text = related.get("message") or "-"
    if related_people:
        related_text = f"{related_text}（命中：{'、'.join([str(x) for x in related_people])}）"

    military_keywords = military.get("keyword_hits") or []
    military_text = military.get("message") or "-"
    if military_keywords:
        military_text = f"{military_text}（关键词：{', '.join([str(x) for x in military_keywords])}）"

    return [
        f"- 来源：{external_checks.get('source') or '-'}",
        f"- 核验时间：{external_checks.get('checked_at') or '-'}",
        f"- 证据文件：{'、'.join([str(x) for x in evidence_files]) if evidence_files else '-'}",
        f"- 甲方名称：{party_external.get('name') or '-'}",
        f"- 统一社会信用代码：{party_external.get('credit_code') or '-'}",
        f"- 法定代表人：{party_external.get('legal_rep') or '-'}",
        f"- 注册地址：{party_external.get('registered_address') or '-'}",
        f"- 经营状态：{party_external.get('company_status') or '-'}",
        f"- 关联关系核验：{related_text}",
        f"- 军工/涉密相关核验：{military_text}",
    ]


def _write_excel_summary(path: Path, review: dict[str, Any], source_hash: str) -> None:
    headers = [
        "source_hash",
        "source_file",
        "decision",
        "confidence",
        "rule_id",
        "scope",
        "severity",
        "title",
        "message",
        "location",
        "evidence",
        "suggestion",
        "generated_at",
    ]
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "总表"
        ws.append(headers)

    findings = review.get("findings", [])
    if not findings:
        ws.append([
            source_hash,
            review.get("source_file", ""),
            review.get("decision", ""),
            review.get("confidence", ""),
            "-",
            "-",
            "-",
            "无问题",
            "-",
            "-",
            "-",
            "-",
            review.get("generated_at", now_iso()),
        ])
    else:
        for f in findings:
            ws.append([
                source_hash,
                review.get("source_file", ""),
                review.get("decision", ""),
                review.get("confidence", ""),
                f.get("rule_id", ""),
                f.get("scope", ""),
                f.get("severity", ""),
                f.get("title", ""),
                f.get("message", ""),
                f.get("location", ""),
                f.get("evidence", ""),
                f.get("suggestion", ""),
                review.get("generated_at", now_iso()),
            ])
    wb.save(path)
