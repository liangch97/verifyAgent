from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_BOOL_TRUE = {"1", "true", "yes", "y", "是", "有", "真"}
_BOOL_FALSE = {"0", "false", "no", "n", "否", "无", "假"}
_LIST_SPLIT_RE = re.compile(r"[、,，;；|\n\r]+")


class CompanyRegistryProvider:
    """合法外部企业信息提供器接口骨架（仅预留，不含爬虫实现）。"""

    def lookup(self, company_name: str) -> dict[str, Any]:
        raise NotImplementedError("请使用合法官方 API 或人工导入结果，不要绕过登录/验证码/付费限制。")


def load_company_check(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"company_check 文件不存在: {path}")

    suffix = path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
    elif suffix == ".csv":
        payload = _load_flat_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        payload = _load_flat_xlsx(path)
    else:
        raise ValueError(f"暂不支持的 company_check 文件类型: {suffix}")

    return normalize_company_check(payload)


def normalize_company_check(payload: dict[str, Any]) -> dict[str, Any]:
    source = _to_text(_pick(payload, "source")) or "人工核验"
    checked_at = _normalize_date(_pick(payload, "checked_at"))
    evidence_files = _to_str_list(_pick(payload, "evidence_files"))

    party_a = {
        "name": _to_text(_pick(payload, "party_a.name")),
        "credit_code": _to_text(_pick(payload, "party_a.credit_code")),
        "legal_rep": _to_text(_pick(payload, "party_a.legal_rep")),
        "registered_address": _to_text(_pick(payload, "party_a.registered_address")),
        "company_status": _to_text(_pick(payload, "party_a.company_status")),
        "company_type": _to_text(_pick(payload, "party_a.company_type")),
        "business_scope": _to_text(_pick(payload, "party_a.business_scope")),
        "shareholders": _to_people_list(_pick(payload, "party_a.shareholders"), meta_key="type"),
        "directors": _to_people_list(_pick(payload, "party_a.directors"), meta_key="title"),
        "executives": _to_people_list(_pick(payload, "party_a.executives"), meta_key="title"),
        "is_military_or_defense_related": _to_bool(_pick(payload, "party_a.is_military_or_defense_related")),
        "military_or_defense_evidence": _to_str_list(_pick(payload, "party_a.military_or_defense_evidence")),
        "related_to_research_team": _to_bool(_pick(payload, "party_a.related_to_research_team")),
        "related_person_matches": _to_str_list(_pick(payload, "party_a.related_person_matches")),
    }

    return {
        "source": source,
        "checked_at": checked_at,
        "evidence_files": evidence_files,
        "party_a": party_a,
        "research_team": _to_research_team(_pick(payload, "research_team")),
    }


def _load_flat_csv(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if any(_to_text(v) for v in row.values()):
                return {(_to_text(k)): v for k, v in row.items() if _to_text(k)}
    return {}


def _load_flat_xlsx(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [_to_text(v) for v in rows[0]]
    for row in rows[1:]:
        if not any(_to_text(v) for v in row):
            continue
        out: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            out[header] = row[idx] if idx < len(row) else ""
        return out
    return {}


def _pick(payload: dict[str, Any], dotted_key: str) -> Any:
    if not isinstance(payload, dict):
        return None
    if dotted_key in payload and payload[dotted_key] not in (None, ""):
        return payload[dotted_key]

    cur: Any = payload
    for part in dotted_key.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
    return cur


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    s = _to_text(value).lower()
    if not s:
        return None
    if s in _BOOL_TRUE:
        return True
    if s in _BOOL_FALSE:
        return False
    return None


def _normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = _to_text(value)
    if not text:
        return datetime.now().strftime("%Y-%m-%d")

    normalized = text.replace("/", "-").replace(".", "-")
    try:
        return datetime.fromisoformat(normalized).strftime("%Y-%m-%d")
    except ValueError:
        return text


def _to_str_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        out: list[str] = []
        for item in value:
            if isinstance(item, dict):
                name = _to_text(item.get("name"))
                if name:
                    out.append(name)
            else:
                text = _to_text(item)
                if text:
                    out.extend(_split_list_text(text))
        return [x for x in out if x]

    if isinstance(value, str):
        parsed = _try_parse_json(value)
        if parsed is not None:
            return _to_str_list(parsed)
        return _split_list_text(value)

    text = _to_text(value)
    return [text] if text else []


def _to_people_list(value: Any, meta_key: str) -> list[dict[str, str]]:
    if value is None:
        return []
    if isinstance(value, str):
        parsed = _try_parse_json(value)
        if parsed is not None:
            return _to_people_list(parsed, meta_key)
        return [{"name": token, meta_key: ""} for token in _split_list_text(value)]
    if isinstance(value, dict):
        name = _to_text(value.get("name"))
        if not name:
            return []
        return [{"name": name, meta_key: _to_text(value.get(meta_key))}]
    if isinstance(value, list):
        out: list[dict[str, str]] = []
        for item in value:
            if isinstance(item, dict):
                name = _to_text(item.get("name"))
                if name:
                    out.append({"name": name, meta_key: _to_text(item.get(meta_key))})
            else:
                text = _to_text(item)
                if text:
                    out.append({"name": text, meta_key: ""})
        return out
    return []


def _to_research_team(value: Any) -> list[dict[str, str]]:
    if value is None:
        return []
    if isinstance(value, str):
        parsed = _try_parse_json(value)
        if parsed is not None:
            return _to_research_team(parsed)
        return [{"name": token, "role": ""} for token in _split_list_text(value)]
    if isinstance(value, dict):
        name = _to_text(value.get("name"))
        if not name:
            return []
        return [{"name": name, "role": _to_text(value.get("role"))}]
    if isinstance(value, list):
        out: list[dict[str, str]] = []
        for item in value:
            if isinstance(item, dict):
                name = _to_text(item.get("name"))
                if name:
                    out.append({"name": name, "role": _to_text(item.get("role"))})
            else:
                text = _to_text(item)
                if text:
                    out.append({"name": text, "role": ""})
        return out
    return []


def _split_list_text(text: str) -> list[str]:
    return [t.strip() for t in _LIST_SPLIT_RE.split(text) if t and t.strip()]


def _try_parse_json(text: str) -> Any | None:
    text = text.strip()
    if not text or text[0] not in {"[", "{"}:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None
